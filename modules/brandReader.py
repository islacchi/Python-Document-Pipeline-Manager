import sys
import re
import os
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

# Dynamic path config resolution
try:
    import config
except ImportError:
    sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    import config

missing = []
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    missing.append("openpyxl")

try:
    from pypdf import PdfReader
    import pdfplumber
except ImportError:
    missing.append("pypdf pdfplumber")

if missing:
    print(f"Missing dependencies. Run: pip install {' '.join(missing)}")
    sys.exit(1)

# ============================================================
#  CONFIGURATION
# ============================================================
MAX_WORKERS    = config.MAX_WORKERS
MAX_PAGES      = config.MAX_PAGES
OCR_DPI        = config.OCR_DPI_HIGH
TEXT_THRESHOLD = config.TEXT_THRESHOLD
TESSERACT_PATH = config.TESSERACT_PATH
POPPLER_PATH   = config.POPPLER_PATH
# ============================================================

try:
    import pytesseract
    from pdf2image import convert_from_path
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False
    print("Warning: OCR not available. Run: pip install pytesseract pdf2image")
    print("         Also install Tesseract: https://github.com/UB-Mannheim/tesseract/wiki\n")

# ── Patterns ─────────────────────────────────────────────────────────────────

BRAND_NAME_PATTERN         = re.compile(r'Brand\s*Name\s*[:\-]', re.IGNORECASE)
BRAND_LABEL_PATTERN        = re.compile(r'Brand\s*Name\s*[:\-\d]\s*([^\n\r:]+)', re.IGNORECASE)
REG_NUMBER_PATTERN         = re.compile(r'Registration\s*Number\s*[:\-]\s*([A-Z0-9\-]+)', re.IGNORECASE)
VALIDITY_PATTERN           = re.compile(r'valid\s+until\s+(\d{1,2}\s+\w+\s+\d{4})', re.IGNORECASE)
MANUFACTURER_PATTERN       = re.compile(r'Manufacturer\s*[:\-\d]\s*([^\n\r:]+)', re.IGNORECASE)
TRADER_PATTERN             = re.compile(r'Trader\s*[:\-\d]\s*([^\n\r:]+)', re.IGNORECASE)
IMPORTER_PATTERN           = re.compile(r'Importer\s*(?:/\s*Distributor)?\s*[:\-\d]\s*([^\n\r:]+)', re.IGNORECASE)
DISTRIBUTOR_PATTERN        = re.compile(r'Distributor\s*[:\-\d]\s*([^\n\r:]+)', re.IGNORECASE)
FDA_REG_NUMBER_PATTERN     = re.compile(r'FDA\s+Registration\s+No\.?\s*[:\-]\s*([A-Z0-9\-]+)', re.IGNORECASE)
MANUFACTURER_NAME_ADDR_PATTERN = re.compile(
    r'Manufacturer\s*(?:Name\s*and\s*Address\s*)?[:\-]?\s*((?:[^\n\r: ]| (?! ))+)',
    re.IGNORECASE
)

# ── Helpers ───────────────────────────────────────────────────────────────────

def natural_sort_key(filename: str):
    return [
        int(part) if part.isdigit() else part.lower()
        for part in re.split(r'(\d+)', filename)
    ]

def _extract_text_pages(pages, extractor_func) -> str:
    return "\n".join(extractor_func(page) or "" for page in pages[:MAX_PAGES])

def extract_text_pypdf(pdf_path: str) -> str:
    reader = PdfReader(pdf_path)
    return _extract_text_pages(reader.pages, lambda p: p.extract_text())

def extract_text_pdfplumber(pdf_path: str) -> str:
    with pdfplumber.open(pdf_path) as pdf:
        return _extract_text_pages(pdf.pages, lambda p: p.extract_text())

def extract_text_ocr(pdf_path: str) -> str:
    images = convert_from_path(
        pdf_path, first_page=1, last_page=MAX_PAGES,
        poppler_path=POPPLER_PATH, dpi=OCR_DPI,
    )
    return "\n".join(pytesseract.image_to_string(img, config="--psm 6") for img in images)

def extract_text_from_pdf(pdf_path: str) -> str:
    text = extract_text_pypdf(pdf_path)
    if len(text.strip()) < TEXT_THRESHOLD or not BRAND_NAME_PATTERN.search(text):
        plumber_text = extract_text_pdfplumber(pdf_path)
        if BRAND_NAME_PATTERN.search(plumber_text) or len(plumber_text.strip()) > len(text.strip()):
            text = plumber_text
    if len(text.strip()) < TEXT_THRESHOLD and OCR_AVAILABLE:
        text = extract_text_ocr(pdf_path)
    return text

def clean_field(raw: str) -> str | None:
    raw = re.split(r'[\n\r\t|]', raw.strip())[0].strip()
    raw = re.sub(r'^[\s:\-]+', '', raw).strip()
    return raw if raw else None

# ── Field extractors ─────────────────────────────────────────────────────────

def extract_brand_name(text: str) -> str | None:
    match = BRAND_LABEL_PATTERN.search(text)
    return clean_field(match.group(1)) if match else None

def extract_registration_number(text: str) -> str | None:
    match = REG_NUMBER_PATTERN.search(text) or FDA_REG_NUMBER_PATTERN.search(text)
    return clean_field(match.group(1)) if match else None

def extract_validity(text: str) -> str | None:
    match = VALIDITY_PATTERN.search(text)
    return match.group(1).strip() if match else None

def extract_manufacturer(text: str) -> str | None:
    match = MANUFACTURER_NAME_ADDR_PATTERN.search(text)
    if match:
        raw = re.sub(r'\s+', ' ', match.group(1)).strip()
        return raw if raw else None
    match = MANUFACTURER_PATTERN.search(text)
    return clean_field(match.group(1)) if match else None

def extract_trader(text: str) -> str | None:
    match = TRADER_PATTERN.search(text)
    return clean_field(match.group(1)) if match else None

def extract_importer(text: str) -> str | None:
    match = IMPORTER_PATTERN.search(text)
    return clean_field(match.group(1)) if match else None

def extract_distributor(text: str) -> str | None:
    match = DISTRIBUTOR_PATTERN.search(text)
    return clean_field(match.group(1)) if match else None

# ── PDF worker ────────────────────────────────────────────────────────────────

def process_pdf(folder_path: str, filename: str) -> tuple:
    pdf_path = os.path.join(folder_path, filename)
    try:
        text = extract_text_from_pdf(pdf_path)
        fields = {
            "brand":        extract_brand_name(text),
            "reg_number":   extract_registration_number(text),
            "validity":     extract_validity(text),
            "manufacturer": extract_manufacturer(text),
            "trader":       extract_trader(text),
            "importer":     extract_importer(text),
            "distributor":  extract_distributor(text),
        }
        return (filename, fields, None)
    except Exception as e:
        return (filename, None, str(e))

# ── Excel writer ──────────────────────────────────────────────────────────────

def write_excel(log_path: str, folder_path: str, pdf_files: list, results: dict) -> None:
    wb = Workbook()

    header_font    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    found_font     = Font(name="Arial", size=10)
    not_found_font = Font(name="Arial", size=10, color="888888", italic=True)
    error_font     = Font(name="Arial", size=10, color="CC0000")
    section_font   = Font(name="Arial", bold=True, size=10)
    brand_font     = Font(name="Arial", size=10, bold=True, color="2E7D32")

    green_fill        = PatternFill("solid", fgColor="2E7D32")
    found_fill        = PatternFill("solid", fgColor="F1F8F1")
    not_found_fill    = PatternFill("solid", fgColor="F9F9F9")
    error_fill        = PatternFill("solid", fgColor="FFF0F0")
    section_fill      = PatternFill("solid", fgColor="E8F5E9")
    white_fill        = PatternFill("solid", fgColor="FFFFFF")
    not_found_section = PatternFill("solid", fgColor="FFF3E0")
    error_section     = PatternFill("solid", fgColor="FFEBEE")

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_cell(cell, font, fill, alignment=left):
        cell.font      = font
        cell.fill      = fill
        cell.alignment = alignment
        cell.border    = border

    ws = wb.active
    ws.title = "Brand Names"
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:M1")
    ws["A1"] = (
        f"Brand Name Extraction  |  Folder: {os.path.abspath(folder_path)}"
        f"  |  Run at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        f"  |  OCR: {'enabled' if OCR_AVAILABLE else 'disabled'}"
    )
    ws["A1"].font      = Font(name="Arial", size=9, italic=True, color="555555")
    ws["A1"].alignment = left
    ws.row_dimensions[1].height = 18

    headers = ["#", "File Name", "Brand Name", "Registration No.", "Valid Until",
               "Manufacturer", "Trader", "Importer", "Distributor"]
    for col, h in enumerate(headers, 1):
        style_cell(ws.cell(row=2, column=col, value=h), header_font, green_fill, center)
    ws.row_dimensions[2].height = 22

    row = 3
    found     = [(f, results[f][0]) for f in pdf_files if results[f][0] and results[f][0].get("brand")]
    not_found = [f for f in pdf_files if results[f][0] and not results[f][0].get("brand") and not results[f][1]]
    errors    = [(f, results[f][1]) for f in pdf_files if results[f][1]]

    if found:
        ws.merge_cells(f"A{row}:I{row}")
        ws[f"A{row}"] = f"✔  FOUND ({len(found)})"
        style_cell(ws[f"A{row}"], section_font, section_fill)
        row += 1
        for i, (filename, fields) in enumerate(found, 1):
            fill = white_fill if i % 2 == 0 else found_fill
            style_cell(ws.cell(row=row, column=1, value=i),                          found_font, fill, center)
            style_cell(ws.cell(row=row, column=2, value=filename),                   found_font, fill)
            style_cell(ws.cell(row=row, column=3, value=fields.get("brand") or "—"), brand_font, fill)
            for col_idx, key in enumerate(["reg_number", "validity", "manufacturer", "trader", "importer", "distributor"], 4):
                v = fields.get(key) or "—"
                style_cell(ws.cell(row=row, column=col_idx, value=v), found_font, fill, center if v == "—" else left)
            row += 1

    if not_found:
        row += 1
        ws.merge_cells(f"A{row}:I{row}")
        ws[f"A{row}"] = f"✘  NOT FOUND ({len(not_found)})"
        style_cell(ws[f"A{row}"], section_font, not_found_section)
        row += 1
        for i, filename in enumerate(not_found, 1):
            fill = not_found_fill if i % 2 != 0 else white_fill
            style_cell(ws.cell(row=row, column=1, value=i),       not_found_font, fill, center)
            style_cell(ws.cell(row=row, column=2, value=filename), not_found_font, fill)
            for col_idx in range(3, 10):
                style_cell(ws.cell(row=row, column=col_idx, value="—"), not_found_font, fill, center)
            row += 1

    if errors:
        row += 1
        ws.merge_cells(f"A{row}:I{row}")
        ws[f"A{row}"] = f"⚠  ERRORS ({len(errors)})"
        style_cell(ws[f"A{row}"], section_font, error_section)
        row += 1
        for i, (filename, err) in enumerate(errors, 1):
            style_cell(ws.cell(row=row, column=1, value=i),       error_font, error_fill, center)
            style_cell(ws.cell(row=row, column=2, value=filename), error_font, error_fill)
            style_cell(ws.cell(row=row, column=3, value=err),      error_font, error_fill)
            for col_idx in range(4, 10):
                style_cell(ws.cell(row=row, column=col_idx, value=""), error_font, error_fill)
            row += 1

    row += 1
    ws.merge_cells(f"A{row}:I{row}")
    ws[f"A{row}"] = f"Total: {len(pdf_files)}   |   Found: {len(found)}   |   Not Found: {len(not_found)}   |   Errors: {len(errors)}"
    ws[f"A{row}"].font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws[f"A{row}"].fill      = green_fill
    ws[f"A{row}"].alignment = center
    ws[f"A{row}"].border    = border

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 45
    ws.column_dimensions["G"].width = 25
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 30

    wb.save(log_path)

# ── Entry point ───────────────────────────────────────────────────────────────

def run(folder_path: str) -> None:
    LOG_FILE = config.BRAND_LOG_FILE
    base, ext = os.path.splitext(LOG_FILE)
    log_path = os.path.join(folder_path, LOG_FILE)
    counter = 1
    while os.path.exists(log_path):
        log_path = os.path.join(folder_path, f"{base}({counter}){ext}")
        counter += 1

    pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".pdf")]
    pdf_files.sort(key=natural_sort_key)

    if not pdf_files:
        print(f"No PDF files found in: {folder_path}")
        return

    print(f"Processing {len(pdf_files)} PDF(s) with {MAX_WORKERS} workers...\n")

    results = {}
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(process_pdf, folder_path, f): f for f in pdf_files}
        for future in as_completed(futures):
            filename, fields, error = future.result()
            results[filename] = (fields, error)
            print(f"{'✓' if fields and fields.get('brand') else '✗'} {filename}")

    write_excel(log_path, folder_path, pdf_files, results)

    found_count = sum(1 for f in pdf_files if results[f][0] and results[f][0].get("brand"))
    print(f"\nDone — {found_count}/{len(pdf_files)} brand names found.")
    print(f"Log saved to: {log_path}")